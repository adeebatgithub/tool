import openpyxl
import sys, os
from os.path import exists
from xl_dupe import dupe_tools
from bfa import *

class xltool:
    
    def __init__(self):
        
        self.file_exists()
            
    def file_exists(self):
        r = exists("/storage/emulated/0/Xltool/")
        if not r:
            os.mkdir("/storage/emulated/0/Xltool/")
            
    def file_inp(self, path=0):
        
        if path == 0:
            path = input_c("[+] File Name      : ")
        try:
            self.book = openpyxl.load_workbook(path)
            self.worksheets = len(self.book.worksheets)
            self.path = path
        
        except FileNotFoundError:
            red(f"[!] file not found : '{path}'")
            if __name__ == "__main__":
                quit()
            self.file_inp()
        except openpyxl.utils.exceptions.InvalidFileException:
            red(f"[!] file not supported : '{path}'")
            if __name__ == "__main__":
                red("exiting...")
                quit()
            self.file_inp()
            
    def sheet_inp(self, sheet_name=0):
        
        if sheet_name == 0:
            sheet_name = input_c("[+] Worksheet Name : ")
        if sheet_name == "-s" or sheet_name == "_sheets_":
            sheet_names = self.book.sheetnames
            print("\033[1;32m[~] ",end="")
            print(*sheet_names,end="",sep="\n[~] ")
            print("\033[0m")
        check_exit(sheet_name)
            
        try:
            self.sheet = self.book[sheet_name]
            self.sheet_name = sheet_name
        except KeyError:
            if sheet_name != "-s":
                print_er(f"worksheet not found : '{sheet_name}'")
            if __name__ == "__main__":
                red("exiting...")
                quit()
            self.sheet_inp()
            
    def content_to_dict(self):
        
        content_dict = {0:[col for col in range(1,len(self.sheet[1])+1)]}
        max_len = 0
        
        for row in self.sheet:
            content_dict[row[0].row] = []
            for cell in row:
                value = cell.value
                if value == None:
                    value = ""
                if len(str(value)) > max_len:
                    max_len = len(value)
                content_dict[cell.row].append(value)
        return content_dict, max_len
        
    def headers(self):
        
        content_dict,temp = self.content_to_dict()    
        head_values = 0
        head_row = 0
        max_len = 0
            
        for row,values in content_dict.items():
            if row == 0: continue
            lst = []
            for value in values:
                if value != "":
                    lst.append(value)
            if len(lst) > max_len:
                head_values = lst
                head_row = row
                max_len = len(lst)
                    
        return head_values, head_row
        
    def save_file(self):

        file_name = get_file_name(self.path)
        path = "/storage/emulated/0/Xltool/"+file_name
        self.book.save(path)
        print("")
        print_ln()
        blue(f"[*] file Saved to : Xltool/{file_name}")
        self.book.close()
                
    def dupe_tool(self, delete=2, header=0, col=0):
        
        tool = dupe_tools(self.sheet)
        blue("[~] checking for duplicates...")
        print()
        if header == 0:
            if col == 0:
                dupe_lst = tool.full_row()
            else:
                dupe_lst = tool.by_col(col)
        else:
            dupe_lst = tool.by_header(header)
        if len(dupe_lst) != 0:
            print()
            blue(f"[~] {len(dupe_lst)} duplicates found")
            print_ln()
            if delete == 2:
                delete = yn_trufls("do you want to delete the row")
                
            if delete:
                count = 0
                for row in dupe_lst:
                    self.sheet.delete_rows(row-count)
                    red(f"[!] Deleted : {row}")
                    count += 1
                self.save_file()
        else:
            green("[~] Duplicates not found ")
                
    def rm_blnk(self, save=True):
        
        row_lst = []
        
        for row in self.sheet:
            count = 0
            for cell in row:
                if cell.value == None:
                    count += 1
            if len(row) == count:
                row_lst.append(row[0].row)
        print("")        
        green("[~] Removing blank rows...")        
        count = 0
        for row in row_lst:
            self.sheet.delete_rows(row-count)
            count += 1
        print()
        green("[~] Blank rows removed")
        if save:
            self.save_file()
            
    def print_sheet(self,mode):
        
        content_dict,max_len = self.content_to_dict()
        
        def lst():
            for row,values in content_dict.items():
                if row == 0: continue
                for value in values:
                    space = max_len-len(str(value))
                    print(f" {value} ",end=" "*space)
                print()
        
        def table():
            for row,values in content_dict.items():
                
                space = len(str(max(content_dict.keys())))-len(str(row))
                print(f" [ {row} ",end=" "*space)
                for value in values:
                    space = int(max_len)-int(len(str(value)))
                    if value == None:
                        value = ""
                    print(f"| {value}",end=" "*space)
                print("]")
        def json():
            
            head_values, head_row = self.headers()
            if head_values != 0:        
                for row, values in content_dict.items():
                    if row == 0: continue
                    if row == head_row: continue
                    print("{ row: "+str(row),end=", ")
                    for head, value in zip(head_values, values):
                        if head == "": continue
                        print("\033[0;33m"+str(head)+"\033[0m :"+str(value).strip(),end=", ")
                    print("}")
            else:
                print_er("can't find headers")
                raise Error()
                    
        mode_dict = {
            "list": lst,
            "table": table,
            "json": json,
        }
        try:
            mode_dict[mode]()
        except KeyError:
            print_er(f"print mode not found : '{mode}'")
            raise
            
    def search_inp(self):
        
        col = input_c("[+] Column : ")
        txt = input_c("[+] Text    : ")
        self.search_ht = col, txt
            
    def search(self, search_txt=0, header=0, col=0):
        content_dict, temp = self.content_to_dict()
        head_values,head_row = self.headers()
        if search_txt == 0:
            col = input_c("[+] Column : ")
            search_txt = input_c("[+] Text   : ")
            print()
        for txt in search_txt.split(","):
            blue(f"[~] Search result for : {txt}")
            print()
            count = 0
            if header == 0 and col == 0:
                for row, values in content_dict.items():
                    for value in values:
                        if str(value) == txt:
                            count += 1
                            print("{ row: "+str(row),end=", ")
                            for head, value in zip(head_values, values):
                                if head == "": continue
                                print("\033[0;33m"+str(head)+"\033[0m: "+str(value).strip(),end=", ")
                            print("}")
                                
            if header != 0 and col == 0:
                head_index = head_values.index(header)
                for row, values in content_dict.items():
                    if str(values[head_index]) == txt:
                        count += 1
                        print("{ row: "+str(row),end=", ")
                        for head, value in zip(head_values, values):
                            if head == "": continue
                            print("\033[0;33m"+str(head)+"\033[0m :"+str(value).strip(),end=", ")
                        print("}")
                        
            if col != 0 and header == 0:
                
                for row in self.sheet:
                    for cell in row:
                        if col.upper() not in cell.coordinate:
                            continue
                        if cell.value == None: continue
                        if str(cell.value) == txt:
                            count += 1
                            values = [cell.value for cell in self.sheet[cell.row]]
                            
                            print("{ row: "+str(row[0].row),end=", ")
                            for head, value in zip(head_values, values):
                                if head == "": continue
                                print("\033[0;33m"+str(head)+"\033[0m: "+str(value).strip(),end=", ")
                            print("}")
                    
            if count == 0:
                blue("[!] No match found")
            else:
                print()
                blue(f"[~] {count} match found")
        
if __name__ == "__main__":
     
    toolbox = xltool()
    
    def dupe(delete=False, header=0, col=0):
        print_ln()
        print()
        if "-y" in sys.argv:
            delete = True
        if "-n" in sys.argv:
            delete = False
        if "-h" in sys.argv:
            header = sys.argv[sys.argv.index("-h")+1]
        if "-c" in sys.argv:
            col = sys.argv[sys.argv.index("-c")+1]
        
        toolbox.dupe_tool(delete, header, col)               
    
    def rm_blnk():  
        save = True   
        if "-no" in sys.argv:     
            save = False 
        toolbox.rm_blnk(save=save)
        
    if len(sys.argv) == 1:
        help_xltools()
        quit()

    if len(sys.argv) == 2:
        if "-h" in sys.argv or "--help" in sys.argv:
            help_xltools()
            quit()
        if "--tool" in sys.argv:
            green("[======= XLTOOLS ======]")
            print_ln()
            print()
            toolbox.file_inp()
            toolbox.chk_sheet()
            option_dict = {
                    "1": dupe,
                    "2": rm_blnk,
                }
            print()                               
            blue("    [1] Remove Duplicates")      
            blue("    [2] Remove blank rows")    
            red("    [x] exit")                  
            print("")                              
            print_ln()
            main_menu(option_dict)
            print()
            quit()
            
    if "-wb" in sys.argv or "--workbook" in sys.argv:
        try:
            path = sys.argv[sys.argv.index("-wb")+1]
        except ValueError:
            path = sys.argv[sys.argv.index("--workbook")+1]
        except IndexError:
            print_er("file not provided")
            quit()
    else:
        help_xltools()
        quit()
    toolbox.file_inp(path=path)
    
    if "-s" in sys.argv or "--sheet" in sys.argv:
        try:
            sheet_name = sys.argv[sys.argv.index("-s")+1]
        except ValueError:
            sheet_name = sys.argv[sys.argv.index("--sheet")+1]
        except IndexError:
            print_er("worksheet not provided")
            quit()
        if sheet_name not in toolbox.book.sheetnames:
            print_er(f"worksheet not found : {sheet_name}")
            green(f"[*] Available worksheets : {toolbox.book.sheetnames}")
            quit()
    else:
        print_er("Worksheet not provided")
        quit()
    toolbox.sheet_inp(sheet_name=sheet_name)
    
    print()
    green("[====== XLTOOLS ======]")
    print()
    print_ln()
    green(f"[*] Workbook Name : {get_file_name(path)}")
    green(f"[*] Worksheet Name : {sheet_name}")
        
    if len(sys.argv) == 5:
        print()
        blue("    [1] Remove Duplicates")
        blue("    [2] Remove blank rows")
        red("    [x] exit")
        print("")
        print_ln()
        print("")
        inp = input_c("[=] Select : ")
        check_exit(inp)
        
        choice_dict = {
            "1": dupe,
            "2": rm_blnk,
        }
        try:
            choice_dict[inp]()
        except KeyError:
            print_er(f"option not found : {inp}")
            
            
    if len(sys.argv) > 5:
        if "-dupe" in sys.argv or "--duplicate" in sys.argv:
            dupe()
        if "--rm-dupe" in sys.argv or "--remove-duplicate" in sys.argv:
            dupe(delete=True)
        if "--rm-blank" in sys.argv or "--removeblank" in sys.argv:
            rm_blnk()
        if "-p" in sys.argv or "--print" in sys.argv:
            try:
                mode = sys.argv[sys.argv.index("-p")+1]
            except ValueError:
                mode = sys.argv[sys.argv.index("--print")+1]
            except IndexError:
                mode = "list"
            if "-" in mode:
                mode = "list"
            print_ln()
            print()
            try:
                toolbox.print_sheet(mode=mode)
            except KeyError:
                quit()
            except Error:
                quit()
            print()
            print_ln()
        if "--print-headers" in sys.argv or "-ph" in sys.argv:
            values,row = toolbox.headers()
            print_ln()
            print("\033[1;32mAvailable headers : \033[0m",end="")
            print(" ",end="")
            print(*values)
        if "--search" in sys.argv or "-sh" in sys.argv:
            try:
                txt = sys.argv[sys.argv.index("-sh")+1]
            except ValueError:
                txt = sys.argv[sys.argv.index("--search")+1]
            except IndexError:
                print_er("search value not provided")
                quit()
            txt_lst = txt.split(",")
            header = 0
            if "-h" in sys.argv:
                header = sys.argv[sys.argv.index("-h")+1]
            col = 0
            if "-c" in sys.argv:
                col = sys.argv[sys.argv.index("-c")+1]
            print_ln()
            toolbox.search(txt,col=col,header=header)
            print_ln()
    
    
